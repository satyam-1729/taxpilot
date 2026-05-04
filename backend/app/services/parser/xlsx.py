"""XLSX utilities — text extraction for the LLM parser.

Brokers (Zerodha, Groww, Upstox, Angel One, etc.) export tax P&L as XLSX.
Rather than rendering the workbook to PDF and sending it through the vision
path, we flatten every sheet to a compact text representation and send that
as a `text` content block — cheaper, faster, more reliable for tabular data.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


# ZIP magic bytes — XLSX (and other Office Open XML formats) are zip files.
XLSX_MAGIC = b"PK\x03\x04"
MAX_ROWS_PER_SHEET = 1500
MAX_COLS_PER_SHEET = 30


def is_xlsx(content: bytes) -> bool:
    """Quick magic-byte sniff for XLSX/zip-format Office files."""
    return content[:4] == XLSX_MAGIC


def xlsx_to_text(content: bytes) -> str:
    """Render every sheet as a markdown-flavoured text dump.

    Empty rows are skipped. Cells are tab-separated. Sheets are headed with
    `## SheetName`. Truncates to 1500 rows × 30 cols per sheet to keep token
    counts predictable.
    """

    # NOTE: read_only=True silently drops rows on some real-world XLSX exports
    # (e.g. Zerodha tax P&L files have a non-standard layout that breaks the
    # streaming reader). Files are capped at 15 MB so loading fully is fine.
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except InvalidFileException as e:
        raise ValueError(f"Not a valid XLSX file: {e}") from e

    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n## Sheet: {sheet_name}\n")
        rows_added = 0
        for row in ws.iter_rows(values_only=True):
            if rows_added >= MAX_ROWS_PER_SHEET:
                parts.append(f"... (truncated at {MAX_ROWS_PER_SHEET} rows)\n")
                break
            cells = [_format_cell(c) for c in row[:MAX_COLS_PER_SHEET]]
            line = "\t".join(cells).rstrip()
            if not line:
                continue
            parts.append(line)
            rows_added += 1
    wb.close()
    return "\n".join(parts).strip()


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Drop trailing zeros while keeping precision.
        if value.is_integer():
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value).strip()
